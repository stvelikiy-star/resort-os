import asyncio
import io
import os
import uuid
from datetime import date, timedelta

import asyncpg
import httpx
from openpyxl import load_workbook

BASE_URL = os.environ.get("RESORT_CORE_TEST_URL", "http://127.0.0.1:8000")
OWNER_USERNAME = os.environ.get("BOOTSTRAP_OWNER_USERNAME", "ci-owner")
OWNER_PASSWORD = os.environ.get("BOOTSTRAP_OWNER_PASSWORD", "CI-Owner-Strong-Password-2026")
DATABASE_URL = os.environ["DATABASE_URL"].split("?")[0]


def choose_option(client: httpx.Client, start: date, end: date, adults: int = 2):
    response = client.get(
        "/api/v1/booking/check-availability",
        params={"check_in": start.isoformat(), "check_out": end.isoformat(), "adults": adults, "children": 0},
    )
    response.raise_for_status()
    return next(item for item in response.json()["results"] if item["available_count"] > 0 and item["pricing"]["sellable"])


def create_reservation(client: httpx.Client, start: date, phone: str, email: str, suffix: str):
    end = start + timedelta(days=2)
    option = choose_option(client, start, end)
    created = client.post(
        "/api/v1/booking/requests",
        json={
            "guest_name": "Repeat Guest CI",
            "phone": phone,
            "email": email,
            "check_in": start.isoformat(),
            "check_out": end.isoformat(),
            "adults": 2,
            "children": 0,
            "room_type_code": option["room_type_code"],
            "source": "CI_OWNER_INTELLIGENCE",
        },
    )
    created.raise_for_status()
    request_id = created.json()["id"]
    quoted = client.post(
        f"/api/v1/admin/booking/requests/{request_id}/quote",
        json={"room_type_code": option["room_type_code"]},
    )
    quoted.raise_for_status()
    confirmed = client.post(
        f"/api/v1/admin/booking/requests/{request_id}/confirm-payment",
        json={
            "amount_kgs": 1200,
            "method": "CI_MANAGER",
            "external_ref": f"owner-intelligence-{suffix}",
            "idempotency_key": f"owner-intelligence-payment-{suffix}",
        },
    )
    confirmed.raise_for_status()
    return confirmed.json(), start, end


async def inject_test_facts():
    conn = await asyncpg.connect(DATABASE_URL)
    main_pid = await conn.fetchval("SELECT id FROM properties WHERE code='THREE_CROWNS'")

    duplicate_phone = "+996700123123"
    duplicate_one, duplicate_two = uuid.uuid4(), uuid.uuid4()
    await conn.execute(
        '''INSERT INTO guests(id,"propertyId","firstName",phone,"createdAt","updatedAt")
           VALUES($1,$2,$3,$4,now(),now()),($5,$2,$6,$4,now(),now())''',
        duplicate_one, main_pid, "Duplicate One", duplicate_phone,
        duplicate_two, "Duplicate Two",
    )

    other_pid, other_guest = uuid.uuid4(), uuid.uuid4()
    await conn.execute(
        'INSERT INTO properties(id,code,name,"createdAt","updatedAt") VALUES($1,$2,$3,now(),now())',
        other_pid, "OWNER_INTELLIGENCE_OTHER", "Other Property",
    )
    await conn.execute(
        'INSERT INTO guests(id,"propertyId","firstName",phone,"createdAt","updatedAt") VALUES($1,$2,$3,$4,now(),now())',
        other_guest, other_pid, "Other Guest", "+996700999999",
    )

    phone_guest, email_guest = uuid.uuid4(), uuid.uuid4()
    await conn.execute(
        '''INSERT INTO guests(id,"propertyId","firstName",phone,email,"createdAt","updatedAt")
           VALUES($1,$2,$3,$4,$5,now(),now()),($6,$2,$7,$8,$9,now(),now())''',
        phone_guest, main_pid, "Phone Owner", "+996700000001", "phone-owner@example.com",
        email_guest, "Email Owner", "+996700000002", "identity-conflict@example.com",
    )
    await conn.close()
    return str(other_guest)


async def prove_conflict_rolled_back(request_id: str):
    conn = await asyncpg.connect(DATABASE_URL)
    reservation_count = await conn.fetchval('SELECT count(*) FROM reservations WHERE "requestId"=$1::uuid', request_id)
    payment_count = await conn.fetchval('SELECT count(*) FROM payments WHERE "requestId"=$1::uuid', request_id)
    request_status = await conn.fetchval('SELECT status::text FROM reservation_requests WHERE id=$1::uuid', request_id)
    await conn.close()
    assert reservation_count == 0
    assert payment_count == 0
    assert request_status == "QUOTED"


def main():
    client = httpx.Client(base_url=BASE_URL, timeout=30.0)
    login = client.post("/api/v1/auth/login", json={"username": OWNER_USERNAME, "password": OWNER_PASSWORD})
    login.raise_for_status()

    first_start = date.today() + timedelta(days=20)
    first, first_in, _ = create_reservation(
        client, first_start, "+996 555 444 333", "repeat.guest@example.com", "001"
    )
    assert first["guest_identity_created"] is True
    assert first["guest_identity_match"] == "NEW"

    second, _, second_out = create_reservation(
        client, first_start + timedelta(days=5), "0555 444 333", "REPEAT.GUEST@EXAMPLE.COM", "002"
    )
    assert second["guest_identity_created"] is False
    assert second["guest_id"] == first["guest_id"]
    assert second["guest_identity_match"] in {"PHONE", "EMAIL", "PHONE_AND_EMAIL"}

    guests = client.get("/api/v1/admin/intelligence/guests", params={"search": "555444333", "limit": 300})
    guests.raise_for_status()
    matching = [item for item in guests.json()["items"] if item["id"] == first["guest_id"]]
    assert len(matching) == 1
    assert matching[0]["reservation_count"] == 2
    assert matching[0]["total_nights"] == 4
    assert matching[0]["received_kgs"] == 2400

    detail = client.get(f"/api/v1/admin/intelligence/guests/{first['guest_id']}")
    detail.raise_for_status()
    detail_body = detail.json()
    assert detail_body["lifetime"]["reservation_count"] == 2
    assert detail_body["lifetime"]["total_nights"] == 4
    assert detail_body["lifetime"]["received_kgs"] == 2400
    assert len(detail_body["reservations"]) == 2
    assert all(item["schedule"] for item in detail_body["reservations"])

    matrix = client.get(
        "/api/v1/admin/intelligence/occupancy-matrix",
        params={"from_date": first_in.isoformat(), "to_date": second_out.isoformat()},
    )
    matrix.raise_for_status()
    matrix_body = matrix.json()
    assert len(matrix_body["rooms"]) == 84
    reservation_segments = [
        segment
        for room in matrix_body["rooms"]
        for segment in room["segments"]
        if segment["kind"] == "RESERVATION"
        and segment["reservation_id"] in {first["reservation_id"], second["reservation_id"]}
    ]
    assert len(reservation_segments) >= 2

    export = client.get(
        "/api/v1/admin/intelligence/export.xlsx",
        params={"from_date": first_in.isoformat(), "to_date": second_out.isoformat()},
    )
    export.raise_for_status()
    assert export.content[:2] == b"PK"
    workbook = load_workbook(io.BytesIO(export.content), read_only=True)
    expected_sheets = {"Итоги", "Занятость по номерам", "Брони", "Гости", "Платежи"}
    assert expected_sheets.issubset(set(workbook.sheetnames))
    assert workbook["Занятость по номерам"].max_row == 85
    workbook.close()

    other_guest_id = asyncio.run(inject_test_facts())

    duplicates = client.get("/api/v1/admin/intelligence/guests/duplicate-candidates")
    duplicates.raise_for_status()
    duplicate_body = duplicates.json()
    assert duplicate_body["automatic_merge"] is False
    assert any(
        group["reason"] == "PHONE" and group["identity_key"] == "996700123123"
        for group in duplicate_body["groups"]
    )

    foreign_guest = client.get(f"/api/v1/admin/intelligence/guests/{other_guest_id}")
    assert foreign_guest.status_code == 404

    conflict_start = first_start + timedelta(days=10)
    conflict_end = conflict_start + timedelta(days=2)
    option = choose_option(client, conflict_start, conflict_end, adults=1)
    request_created = client.post(
        "/api/v1/booking/requests",
        json={
            "guest_name": "Identity Conflict CI",
            "phone": "+996700000001",
            "email": "identity-conflict@example.com",
            "check_in": conflict_start.isoformat(),
            "check_out": conflict_end.isoformat(),
            "adults": 1,
            "children": 0,
            "room_type_code": option["room_type_code"],
            "source": "CI_OWNER_CONFLICT",
        },
    )
    request_created.raise_for_status()
    conflict_request_id = request_created.json()["id"]
    quote = client.post(
        f"/api/v1/admin/booking/requests/{conflict_request_id}/quote",
        json={"room_type_code": option["room_type_code"]},
    )
    quote.raise_for_status()
    conflict = client.post(
        f"/api/v1/admin/booking/requests/{conflict_request_id}/confirm-payment",
        json={
            "amount_kgs": 1000,
            "method": "CI_MANAGER",
            "external_ref": "owner-intelligence-conflict",
            "idempotency_key": "owner-intelligence-payment-conflict",
        },
    )
    assert conflict.status_code == 409
    assert conflict.json()["detail"]["code"] == "GUEST_IDENTITY_CONFLICT"
    asyncio.run(prove_conflict_rolled_back(conflict_request_id))

    client.close()
    print("OWNER_INTELLIGENCE_E2E_OK")


if __name__ == "__main__":
    main()
