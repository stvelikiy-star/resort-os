from __future__ import annotations

import uuid
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from .auth import require_roles

router = APIRouter(prefix="/api/v1/admin/intelligence", tags=["owner-intelligence"])
manager_access = require_roles("OWNER", "MANAGER")


async def property_context(conn, property_code: str):
    row = await conn.fetchrow(
        'SELECT id,code,name,timezone,currency FROM properties WHERE code=$1',
        property_code,
    )
    if not row:
        raise HTTPException(status_code=503, detail="Property not loaded")
    return row


def validate_range(from_date: date, to_date: date, max_days: int = 367) -> tuple[date, int]:
    if to_date < from_date:
        raise HTTPException(status_code=422, detail="to_date must be on or after from_date")
    days = (to_date - from_date).days + 1
    if days > max_days:
        raise HTTPException(status_code=422, detail=f"range cannot exceed {max_days} calendar days")
    return to_date + timedelta(days=1), days


def guest_display_name(row) -> str:
    parts = [row.get("firstName"), row.get("lastName")]
    value = " ".join(str(part).strip() for part in parts if part and str(part).strip())
    return value or "Гость"


async def fetch_guest_rows(conn, property_id, search: str | None, limit: int, offset: int):
    needle = (search or "").strip()
    return await conn.fetch(
        '''
        SELECT g.id,g."firstName",g."lastName",g.phone,g.email,g."createdAt",g."updatedAt",
               COALESCE(stats.reservation_count,0)::int AS reservation_count,
               COALESCE(stats.completed_stays,0)::int AS completed_stays,
               COALESCE(stats.total_nights,0)::int AS total_nights,
               COALESCE(stats.booked_value_kgs,0)::bigint AS booked_value_kgs,
               COALESCE(stats.received_kgs,0)::bigint AS received_kgs,
               stats.last_stay,stats.next_stay,stats.latest_source
        FROM guests g
        LEFT JOIN LATERAL (
          SELECT
            count(r.id)::int AS reservation_count,
            count(r.id) FILTER (WHERE r.status='CHECKED_OUT')::int AS completed_stays,
            COALESCE(SUM((r."checkOut"-r."checkIn")) FILTER (
              WHERE r.status IN ('GUARANTEED','CHECKED_IN','CHECKED_OUT')
            ),0)::int AS total_nights,
            COALESCE(SUM(r."totalKgs") FILTER (
              WHERE r.status IN ('GUARANTEED','CHECKED_IN','CHECKED_OUT')
            ),0)::bigint AS booked_value_kgs,
            COALESCE((
              SELECT SUM(p."amountKgs")
              FROM reservations rp
              JOIN payments p ON p."reservationId"=rp.id AND p.status='RECEIVED'
              WHERE rp."primaryGuestId"=g.id
            ),0)::bigint AS received_kgs,
            MAX(r."checkOut") FILTER (WHERE r.status='CHECKED_OUT') AS last_stay,
            MIN(r."checkIn") FILTER (WHERE r.status='GUARANTEED') AS next_stay,
            (ARRAY_AGG(rr.source ORDER BY r."createdAt" DESC) FILTER (WHERE rr.source IS NOT NULL))[1] AS latest_source
          FROM reservations r
          LEFT JOIN reservation_requests rr ON rr.id=r."requestId"
          WHERE r."primaryGuestId"=g.id
        ) stats ON true
        WHERE g."propertyId"=$1
          AND (
            $2::text='' OR
            COALESCE(g."firstName",'') ILIKE '%'||$2||'%' OR
            COALESCE(g."lastName",'') ILIKE '%'||$2||'%' OR
            COALESCE(g.phone,'') ILIKE '%'||$2||'%' OR
            COALESCE(g.email,'') ILIKE '%'||$2||'%'
          )
        ORDER BY COALESCE(stats.next_stay,stats.last_stay,g."createdAt"::date) DESC,g."updatedAt" DESC
        LIMIT $3 OFFSET $4
        ''',
        property_id,
        needle,
        limit,
        offset,
    )


async def build_occupancy_matrix(conn, property_id, from_date: date, end_exclusive: date):
    rooms = await conn.fetch(
        '''
        SELECT r.id,r.code,r.name,r."buildingOrZone",r."floorLabel",r."operationalState"::text AS operational_state,
               rt.code AS room_type_code,rt.name AS room_type_name
        FROM rooms r
        JOIN room_types rt ON rt.id=r."roomTypeId"
        WHERE r."propertyId"=$1
        ORDER BY COALESCE(r."buildingOrZone",''),COALESCE(r."floorLabel",''),r.code
        ''',
        property_id,
    )
    blocks = await conn.fetch(
        '''
        SELECT ib.id,ib."roomId",ib."blockType"::text AS block_type,ib."startDate",ib."endDate",ib.reason,
               res.id AS reservation_id,res."bookingNumber",res.status::text AS reservation_status,
               g."firstName",g."lastName"
        FROM inventory_blocks ib
        JOIN rooms room ON room.id=ib."roomId"
        LEFT JOIN reservations res ON res.id=ib."reservationId"
        LEFT JOIN guests g ON g.id=res."primaryGuestId"
        WHERE room."propertyId"=$1
          AND ib.active=true
          AND ib."endDate">$2::date
          AND ib."startDate"<$3::date
          AND (ib."blockType"<>'RESERVATION' OR res.status IN ('GUARANTEED','CHECKED_IN','CHECKED_OUT'))
        ORDER BY room.code,ib."startDate",ib."endDate"
        ''',
        property_id,
        from_date,
        end_exclusive,
    )
    by_room: dict[uuid.UUID, list[dict[str, Any]]] = defaultdict(list)
    for row in blocks:
        by_room[row["roomId"]].append(
            {
                "id": str(row["id"]),
                "kind": row["block_type"],
                "start": max(row["startDate"], from_date),
                "end": min(row["endDate"], end_exclusive),
                "reason": row["reason"],
                "reservation_id": str(row["reservation_id"]) if row["reservation_id"] else None,
                "booking_number": row["bookingNumber"],
                "reservation_status": row["reservation_status"],
                "guest_name": guest_display_name(row) if row["reservation_id"] else None,
            }
        )
    return [
        {
            "id": str(room["id"]),
            "code": room["code"],
            "name": room["name"],
            "building": room["buildingOrZone"],
            "floor": room["floorLabel"],
            "operational_state": room["operational_state"],
            "room_type_code": room["room_type_code"],
            "room_type_name": room["room_type_name"],
            "segments": by_room.get(room["id"], []),
        }
        for room in rooms
    ]


@router.get("/guests")
async def guests(
    request: Request,
    search: str | None = Query(default=None, max_length=160),
    limit: int = Query(default=100, ge=1, le=300),
    offset: int = Query(default=0, ge=0),
    user: dict[str, Any] = Depends(manager_access),
):
    async with request.app.state.db.acquire() as conn:
        prop = await property_context(conn, user["property_code"])
        total = await conn.fetchval(
            '''SELECT count(*) FROM guests WHERE "propertyId"=$1''',
            prop["id"],
        )
        rows = await fetch_guest_rows(conn, prop["id"], search, limit, offset)
    return {
        "items": [
            {
                "id": str(row["id"]),
                "name": guest_display_name(row),
                "phone": row["phone"],
                "email": row["email"],
                "reservation_count": row["reservation_count"],
                "completed_stays": row["completed_stays"],
                "total_nights": row["total_nights"],
                "booked_value_kgs": int(row["booked_value_kgs"] or 0),
                "received_kgs": int(row["received_kgs"] or 0),
                "last_stay": row["last_stay"],
                "next_stay": row["next_stay"],
                "latest_source": row["latest_source"],
                "created_at": row["createdAt"],
                "updated_at": row["updatedAt"],
            }
            for row in rows
        ],
        "total_profiles": int(total or 0),
        "offset": offset,
        "limit": limit,
        "truth": "Guest lifetime values are management metrics from stored Reservation and RECEIVED Payment facts; they are not statutory accounting.",
    }


@router.get("/guests/duplicate-candidates")
async def duplicate_candidates(
    request: Request,
    user: dict[str, Any] = Depends(manager_access),
):
    async with request.app.state.db.acquire() as conn:
        prop = await property_context(conn, user["property_code"])
        rows = await conn.fetch(
            '''
            WITH identities AS (
              SELECT id,"firstName","lastName",phone,email,
                     NULLIF(regexp_replace(COALESCE(phone,''),'\\D','','g'),'') AS phone_key,
                     NULLIF(lower(trim(COALESCE(email,''))),'') AS email_key
              FROM guests WHERE "propertyId"=$1
            ), duplicate_ids AS (
              SELECT id,'PHONE'::text AS reason,phone_key AS identity_key
              FROM identities i
              WHERE phone_key IS NOT NULL AND EXISTS (
                SELECT 1 FROM identities j WHERE j.phone_key=i.phone_key AND j.id<>i.id
              )
              UNION ALL
              SELECT id,'EMAIL'::text AS reason,email_key AS identity_key
              FROM identities i
              WHERE email_key IS NOT NULL AND EXISTS (
                SELECT 1 FROM identities j WHERE j.email_key=i.email_key AND j.id<>i.id
              )
            )
            SELECT d.reason,d.identity_key,i.id,i."firstName",i."lastName",i.phone,i.email
            FROM duplicate_ids d JOIN identities i ON i.id=d.id
            ORDER BY d.reason,d.identity_key,i.id
            ''',
            prop["id"],
        )
    groups: dict[str, dict[str, Any]] = {}
    for row in rows:
        key = f"{row['reason']}:{row['identity_key']}"
        group = groups.setdefault(key, {"reason": row["reason"], "identity_key": row["identity_key"], "guests": []})
        group["guests"].append(
            {"id": str(row["id"]), "name": guest_display_name(row), "phone": row["phone"], "email": row["email"]}
        )
    return {
        "groups": list(groups.values()),
        "automatic_merge": False,
        "truth": "Candidates are evidence for manual review only. Resort Core does not silently merge existing guest histories.",
    }


@router.get("/guests/{guest_id}")
async def guest_detail(
    guest_id: uuid.UUID,
    request: Request,
    user: dict[str, Any] = Depends(manager_access),
):
    async with request.app.state.db.acquire() as conn:
        prop = await property_context(conn, user["property_code"])
        guest = await conn.fetchrow(
            '''SELECT id,"firstName","lastName",phone,email,"createdAt","updatedAt" FROM guests WHERE id=$1 AND "propertyId"=$2''',
            guest_id,
            prop["id"],
        )
        if not guest:
            raise HTTPException(status_code=404, detail="Guest not found")
        reservations = await conn.fetch(
            '''
            WITH paid AS (
              SELECT "reservationId",COALESCE(SUM("amountKgs") FILTER (WHERE status='RECEIVED'),0)::bigint AS paid_kgs
              FROM payments WHERE "reservationId" IS NOT NULL GROUP BY "reservationId"
            )
            SELECT r.id,r."bookingNumber",r.status::text AS status,r."checkIn",r."checkOut",r.adults,r.children,
                   r."totalKgs",r.notes,r."createdAt",r."updatedAt",r."requestId",
                   COALESCE(p.paid_kgs,0)::bigint AS paid_kgs,rr.source
            FROM reservations r
            LEFT JOIN paid p ON p."reservationId"=r.id
            LEFT JOIN reservation_requests rr ON rr.id=r."requestId"
            WHERE r."propertyId"=$1 AND r."primaryGuestId"=$2
            ORDER BY r."checkIn" DESC,r."createdAt" DESC
            ''',
            prop["id"],
            guest_id,
        )
        reservation_ids = [row["id"] for row in reservations]
        request_ids = [row["requestId"] for row in reservations if row["requestId"]]
        schedules = []
        services = []
        payments = []
        conversations = []
        if reservation_ids:
            schedules = await conn.fetch(
                '''
                SELECT ib."reservationId",ib."startDate",ib."endDate",room.code AS room_code,rt.name AS room_type_name
                FROM inventory_blocks ib
                JOIN rooms room ON room.id=ib."roomId"
                JOIN room_types rt ON rt.id=room."roomTypeId"
                WHERE ib."reservationId"=ANY($1::uuid[]) AND ib."blockType"='RESERVATION'
                ORDER BY ib."startDate",room.code
                ''',
                reservation_ids,
            )
            services = await conn.fetch(
                '''
                SELECT id,"reservationId","serviceCode","serviceDate","serviceTime",status::text AS status,
                       priority::text AS priority,title,description,"createdAt","completedAt"
                FROM operational_tasks
                WHERE "propertyId"=$1 AND "reservationId"=ANY($2::uuid[]) AND type='GUEST_REQUEST'
                ORDER BY "createdAt" DESC
                ''',
                prop["id"],
                reservation_ids,
            )
            payments = await conn.fetch(
                '''
                SELECT id,"reservationId","amountKgs",method,status::text AS status,provider,"externalRef","paidAt","createdAt"
                FROM payments WHERE "reservationId"=ANY($1::uuid[]) ORDER BY COALESCE("paidAt","createdAt") DESC
                ''',
                reservation_ids,
            )
        if request_ids:
            conversations = await conn.fetch(
                '''
                SELECT c.id,c.status::text AS status,c."reservationRequestId",c."contactName",c."contactPhone",
                       ch.kind::text AS channel_kind,ch."displayName" AS channel_name,
                       c."lastInboundAt",c."lastOutboundAt",c."createdAt",c."updatedAt",
                       count(m.id)::int AS message_count
                FROM conversations c
                JOIN communication_channels ch ON ch.id=c."channelId"
                LEFT JOIN conversation_messages m ON m."conversationId"=c.id
                WHERE c."propertyId"=$1 AND c."reservationRequestId"=ANY($2::uuid[])
                GROUP BY c.id,ch.kind,ch."displayName"
                ORDER BY c."updatedAt" DESC
                ''',
                prop["id"],
                request_ids,
            )

    schedules_by: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in schedules:
        schedules_by[str(row["reservationId"])].append(
            {"start": row["startDate"], "end": row["endDate"], "room_code": row["room_code"], "room_type_name": row["room_type_name"]}
        )
    services_by: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in services:
        services_by[str(row["reservationId"])].append(
            {
                "id": str(row["id"]), "service_code": row["serviceCode"], "service_date": row["serviceDate"],
                "service_time": row["serviceTime"], "status": row["status"], "priority": row["priority"],
                "title": row["title"], "description": row["description"], "created_at": row["createdAt"], "completed_at": row["completedAt"],
            }
        )
    payments_by: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in payments:
        payments_by[str(row["reservationId"])].append(
            {
                "id": str(row["id"]), "amount_kgs": int(row["amountKgs"]), "method": row["method"],
                "status": row["status"], "provider": row["provider"], "external_ref": row["externalRef"],
                "paid_at": row["paidAt"], "created_at": row["createdAt"],
            }
        )
    total_booked = sum(int(row["totalKgs"]) for row in reservations if row["status"] in {"GUARANTEED", "CHECKED_IN", "CHECKED_OUT"})
    total_received = sum(int(row["amountKgs"]) for row in payments if row["status"] == "RECEIVED")
    total_nights = sum(max((row["checkOut"] - row["checkIn"]).days, 0) for row in reservations if row["status"] in {"GUARANTEED", "CHECKED_IN", "CHECKED_OUT"})

    return {
        "guest": {
            "id": str(guest["id"]), "name": guest_display_name(guest), "phone": guest["phone"], "email": guest["email"],
            "created_at": guest["createdAt"], "updated_at": guest["updatedAt"],
        },
        "lifetime": {
            "reservation_count": len(reservations),
            "completed_stays": sum(1 for row in reservations if row["status"] == "CHECKED_OUT"),
            "total_nights": total_nights,
            "booked_value_kgs": total_booked,
            "received_kgs": total_received,
        },
        "reservations": [
            {
                "id": str(row["id"]), "booking_number": row["bookingNumber"], "status": row["status"],
                "check_in": row["checkIn"], "check_out": row["checkOut"], "adults": row["adults"], "children": row["children"],
                "total_kgs": int(row["totalKgs"]), "paid_kgs": int(row["paid_kgs"] or 0),
                "outstanding_kgs": max(int(row["totalKgs"]) - int(row["paid_kgs"] or 0), 0),
                "source": row["source"], "notes": row["notes"], "created_at": row["createdAt"], "updated_at": row["updatedAt"],
                "schedule": schedules_by[str(row["id"])],
                "services": services_by[str(row["id"])],
                "payments": payments_by[str(row["id"])],
            }
            for row in reservations
        ],
        "conversations": [
            {
                "id": str(row["id"]), "status": row["status"], "request_id": str(row["reservationRequestId"]) if row["reservationRequestId"] else None,
                "channel_kind": row["channel_kind"], "channel_name": row["channel_name"], "message_count": row["message_count"],
                "last_inbound_at": row["lastInboundAt"], "last_outbound_at": row["lastOutboundAt"], "created_at": row["createdAt"], "updated_at": row["updatedAt"],
            }
            for row in conversations
        ],
        "truth": "History is composed only from stored Resort Core guest, reservation, inventory, payment, guest-service and conversation facts.",
    }


@router.get("/occupancy-matrix")
async def occupancy_matrix(
    request: Request,
    from_date: date = Query(),
    to_date: date = Query(),
    user: dict[str, Any] = Depends(manager_access),
):
    end_exclusive, days = validate_range(from_date, to_date, max_days=93)
    async with request.app.state.db.acquire() as conn:
        prop = await property_context(conn, user["property_code"])
        rooms = await build_occupancy_matrix(conn, prop["id"], from_date, end_exclusive)
    return {
        "range": {"from": from_date, "to": to_date, "days": days},
        "dates": [from_date + timedelta(days=index) for index in range(days)],
        "rooms": rooms,
        "truth": "Reservation cells reflect active reservation inventory blocks and current stored reservation status. Maintenance/manual blocks are shown separately.",
    }


@router.get("/export.xlsx")
async def export_xlsx(
    request: Request,
    from_date: date = Query(),
    to_date: date = Query(),
    user: dict[str, Any] = Depends(manager_access),
):
    end_exclusive, days = validate_range(from_date, to_date, max_days=367)
    async with request.app.state.db.acquire() as conn:
        prop = await property_context(conn, user["property_code"])
        rooms = await build_occupancy_matrix(conn, prop["id"], from_date, end_exclusive)
        reservations = await conn.fetch(
            '''
            WITH paid AS (
              SELECT "reservationId",COALESCE(SUM("amountKgs") FILTER (WHERE status='RECEIVED'),0)::bigint AS paid_kgs
              FROM payments WHERE "reservationId" IS NOT NULL GROUP BY "reservationId"
            )
            SELECT r.id,r."bookingNumber",r.status::text AS status,r."checkIn",r."checkOut",r.adults,r.children,r."totalKgs",
                   g."firstName",g."lastName",g.phone,g.email,COALESCE(p.paid_kgs,0)::bigint AS paid_kgs,rr.source
            FROM reservations r
            LEFT JOIN guests g ON g.id=r."primaryGuestId"
            LEFT JOIN paid p ON p."reservationId"=r.id
            LEFT JOIN reservation_requests rr ON rr.id=r."requestId"
            WHERE r."propertyId"=$1 AND r."checkOut">$2::date AND r."checkIn"<$3::date
            ORDER BY r."checkIn",r."bookingNumber"
            ''',
            prop["id"], from_date, end_exclusive,
        )
        guest_rows = await fetch_guest_rows(conn, prop["id"], None, 10000, 0)
        payments = await conn.fetch(
            '''
            SELECT p.id,p."reservationId",r."bookingNumber",p."amountKgs",p.method,p.status::text AS status,p.provider,p."externalRef",p."paidAt",p."createdAt"
            FROM payments p
            LEFT JOIN reservations r ON r.id=p."reservationId"
            LEFT JOIN reservation_requests rr ON rr.id=p."requestId"
            WHERE COALESCE(r."propertyId",rr."propertyId")=$1
              AND (COALESCE(p."paidAt",p."createdAt") AT TIME ZONE $4)::date >= $2::date
              AND (COALESCE(p."paidAt",p."createdAt") AT TIME ZONE $4)::date < $3::date
            ORDER BY COALESCE(p."paidAt",p."createdAt")
            ''',
            prop["id"], from_date, end_exclusive, prop["timezone"],
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Итоги"
    summary = [
        ("Отель", prop["name"]),
        ("Период", f"{from_date.isoformat()} — {to_date.isoformat()}"),
        ("Дней", days),
        ("Номеров", len(rooms)),
        ("Броней в периоде", len(reservations)),
        ("Профилей гостей", len(guest_rows)),
        ("Получено оплат в периоде, KGS", sum(int(row["amountKgs"]) for row in payments if row["status"] == "RECEIVED")),
        ("Сформировано", datetime.utcnow().isoformat(timespec="seconds") + "Z"),
        ("Примечание", "Управленческий отчёт Resort Core; не является бухгалтерской/налоговой отчётностью."),
    ]
    for key, value in summary:
        ws.append([key, value])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 70
    for cell in ws[1]:
        cell.font = Font(bold=True)

    occupancy = wb.create_sheet("Занятость по номерам")
    dates = [from_date + timedelta(days=index) for index in range(days)]
    occupancy.append(["Номер", "Категория", "Корпус", "Этаж", *[d.isoformat() for d in dates]])
    for room in rooms:
        values = []
        for day in dates:
            value = ""
            for segment in room["segments"]:
                if segment["start"] <= day < segment["end"]:
                    if segment["kind"] == "RESERVATION":
                        value = f"{segment['booking_number'] or 'Бронь'} · {segment['guest_name'] or 'Гость'}"
                    else:
                        value = segment["kind"]
                    break
            values.append(value)
        occupancy.append([room["code"], room["room_type_name"], room["building"], room["floor"], *values])
    occupancy.freeze_panes = "E2"
    occupancy.auto_filter.ref = occupancy.dimensions

    rs = wb.create_sheet("Брони")
    rs.append(["Booking", "Статус", "Гость", "Телефон", "Email", "Заезд", "Выезд", "Ночей", "Взрослые", "Дети", "Стоимость KGS", "Оплачено KGS", "Остаток KGS", "Источник"])
    for row in reservations:
        paid = int(row["paid_kgs"] or 0)
        total = int(row["totalKgs"])
        rs.append([
            row["bookingNumber"], row["status"], guest_display_name(row), row["phone"], row["email"],
            row["checkIn"].isoformat(), row["checkOut"].isoformat(), max((row["checkOut"]-row["checkIn"]).days, 0),
            row["adults"], row["children"], total, paid, max(total-paid, 0), row["source"],
        ])
    rs.freeze_panes = "A2"
    rs.auto_filter.ref = rs.dimensions

    gs = wb.create_sheet("Гости")
    gs.append(["Guest ID", "Имя", "Телефон", "Email", "Броней", "Проживаний", "Ночей", "Стоимость броней KGS", "Получено KGS", "Последний выезд", "Следующий заезд", "Источник"])
    for row in guest_rows:
        gs.append([
            str(row["id"]), guest_display_name(row), row["phone"], row["email"], row["reservation_count"], row["completed_stays"],
            row["total_nights"], int(row["booked_value_kgs"] or 0), int(row["received_kgs"] or 0),
            row["last_stay"].isoformat() if row["last_stay"] else None, row["next_stay"].isoformat() if row["next_stay"] else None, row["latest_source"],
        ])
    gs.freeze_panes = "A2"
    gs.auto_filter.ref = gs.dimensions

    ps = wb.create_sheet("Платежи")
    ps.append(["Payment ID", "Booking", "Сумма KGS", "Метод", "Статус", "Provider", "Внешняя ссылка", "Оплачено", "Создано"])
    for row in payments:
        ps.append([
            str(row["id"]), row["bookingNumber"], int(row["amountKgs"]), row["method"], row["status"], row["provider"], row["externalRef"],
            row["paidAt"].isoformat() if row["paidAt"] else None, row["createdAt"].isoformat() if row["createdAt"] else None,
        ])
    ps.freeze_panes = "A2"
    ps.auto_filter.ref = ps.dimensions

    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for column in sheet.columns:
            letter = column[0].column_letter
            width = min(max((len(str(cell.value)) if cell.value is not None else 0) for cell in column) + 2, 42)
            sheet.column_dimensions[letter].width = max(width, 10)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"three-crowns-owner-report-{from_date.isoformat()}-{to_date.isoformat()}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
